import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import win32com.client as win32
import re

def generate_daily_schedule_report(input_file_path):
    """
    Main function to generate the daily schedule report from an Excel file.
    
    Parameters:
    input_file_path (str): Path to the input Excel file
    """
    today = datetime.now().strftime("%d%m%Y")
    output_excel_path = f"Schedule_Report_{today}.xlsx"
    output_word_path = f"Schedule_Report_{today}.docx"
    
    print(f"Starting to process file: {input_file_path}")
    
    # Step 1: Open the Excel document
    try:
        # Read the Excel file with header in row 3
        df = pd.read_excel(input_file_path, header=2)
        print("Excel file loaded successfully")
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return
    
    # Step 2: Fix header alignment (this will be handled during the Excel file writing)
    
    # Step 3: Delete unnecessary columns
    columns_to_drop = [
        'Trip Accepted', 'Vehicle Type', 'null', 'Date', 'Trip Direction', 
        'Miles', 'Fare', 'Group#', 'Trip Type', 'StandingOrder Id', 'One Way'
    ]
    
    # Drop only columns that exist in the dataframe
    columns_to_drop = [col for col in columns_to_drop if col in df.columns]
    df = df.drop(columns=columns_to_drop)
    print("Dropped unnecessary columns")
    
    # Step 4: Add "*Monitored" to Names based on Comments
    df['Name'] = df.apply(
        lambda row: f"{row['Name']} *Monitored" 
        if isinstance(row.get('Comments', ''), str) and 
        (re.search(r'MONITOR', row['Comments'], re.IGNORECASE) or 
         re.search(r'MT', row['Comments'], re.IGNORECASE))
        else row['Name'],
        axis=1
    )
    print("Added *Monitored to names based on comments")
    
    # Step 5: Delete Comments Column
    if 'Comments' in df.columns:
        df = df.drop(columns=['Comments'])
        print("Deleted Comments column")
    
    # Step 6: Group identical names and assign group numbers
    # Create a dictionary to map names to their first occurrence index
    name_indices = {}
    group_numbers = []
    
    for idx, name in enumerate(df['Name']):
        if name not in name_indices:
            name_indices[name] = len(name_indices) + 1
            group_numbers.append(name_indices[name])
        else:
            group_numbers.append(np.nan)  # Only the first name gets a number
    
    # Add the group number column
    df.insert(0, 'Group Number', group_numbers)
    print("Added group numbers for identical names")
    
    # Save the processed dataframe to Excel
    df.to_excel(output_excel_path, index=False)
    print(f"Saved processed data to {output_excel_path}")
    
    # Steps 7-11: Format the Excel file using openpyxl
    format_excel_file(output_excel_path)
    print("Formatted Excel file")
    
    # Step 12: Copy and paste to Word document in landscape layout
    excel_to_word(output_excel_path, output_word_path)
    print(f"Created Word document: {output_word_path}")
    
    print("Process completed successfully!")
    return output_word_path

def format_excel_file(file_path):
    """
    Format the Excel file with specific styles
    
    Parameters:
    file_path (str): Path to the Excel file to format
    """
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply formatting to all cells in the used range
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            # Add borders
            cell.border = thin_border
            
            # Change font to Aptos size 11
            cell.font = Font(name='Aptos', size=11)
            
            # Set alignment
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Format header row
    for cell in ws[1]:
        cell.font = Font(name='Aptos', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Change layout spacing to 0 (this is handled by alignment properties)
    
    # Resize columns to fit content
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 30)  # Cap width at 30
    
    wb.save(file_path)

def excel_to_word(excel_path, word_path):
    """
    Copy the Excel content to a Word document in landscape layout
    
    Parameters:
    excel_path (str): Path to the source Excel file
    word_path (str): Path for the output Word document
    """
    try:
        # Create Word and Excel application objects
        word = win32.Dispatch('Word.Application')
        word.Visible = False
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False
        
        # Open Excel workbook and select the first worksheet
        workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
        worksheet = workbook.Worksheets(1)
        
        # Copy the range
        used_range = worksheet.UsedRange
        used_range.Copy()
        
        # Create a new Word document
        doc = word.Documents.Add()
        
        # Set orientation to landscape
        doc.PageSetup.Orientation = 1  # 0=Portrait, 1=Landscape
        
        # Paste the Excel data into Word
        word.Selection.PasteExcelTable(False, False, False)
        
        # Save and close the Word document
        doc.SaveAs(os.path.abspath(word_path))
        doc.Close()
        
        # Close Excel without saving changes
        workbook.Close(False)
        
        # Quit applications
        word.Quit()
        excel.Quit()
        
    except Exception as e:
        print(f"Error during Excel to Word conversion: {e}")
        
        # Make sure to quit the applications even if there's an error
        try:
            word.Quit()
            excel.Quit()
        except:
            pass

if __name__ == "__main__":
    # Set the input file path - replace with your actual file path
    input_file = "C:\\Users\\user1\\Documents\\PythonRPA\\DadScheduler\\rawdata.xlsx"  # Modify this to your actual file path
    
    # Generate the report
    output_file = generate_daily_schedule_report(input_file)
    print(f"Report generated successfully: {output_file}")